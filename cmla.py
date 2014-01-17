import sys
import getopt
import cmlaTeam as cmla
import random
import tkinter
import math
from types import *
from tkinter import messagebox
from tkinter import filedialog
from pandas.io.parsers import ExcelFile
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook


#
#   make a table of teams
#   initialized to all zeroes
#   Note that team numbers are indexed
#   to zero 

def getScheduleTable(numTeams):
    """
    getScheduleTable will read in the scheduling table from 
    a text file.  The scheduling table determines which teams 
    play each other in a 10 game season (the standard CMLA season)

        int numTeams    the number of teams in the division

        Return value:   list    scheduleTable   a list of tuples determining which 
                                                teams play each other, e.g., (1,2)
                                                Team 1 plays Team 2
    """
    
    #
    # Note the file name is set as [numTeams]teams
    # e.g. if numTeams is 20, then the filename is
    # 20teams

    Filename = str(numTeams) + "teams"

    try:
        if len(scheduleTable) > 0:
            scheduleTable.clear()
    except Exception:
        scheduleTable = []

    linenum = 0
    for line in open(Filename,"r"):
        # print (line)
        tokens = line.split()
        linenum = linenum + 1
#        print (tokens)
        if len(tokens) > 1:
            if tokens[0] == "#":
                # comment line ignore
                print ('Comment ',line)
                pass
            else:
#                print (tokens)
                if len(tokens) == 3:
                    team1 = int(tokens[0]) - 1
                    team2 = int(tokens[2]) - 1
                    scheduleTable.append((team1, team2))
    return scheduleTable


def readRegistrationExcelFile(filename):
    #
    # open the file and set the active worksheet
    wb = load_workbook(filename)
    ws = wb.get_active_sheet()
    #xd = pd.ExcelFile(filename)
    #df = xd.parse('Sheet1')
    #df.fillna(0,inplace=True)
    #parishList = list(df.columns[2:])
    #gradeList = list(df.iloc[0:,0])
    #genderList = list(df.iloc[0:,1])
    parishList = []
    gradeList = []
    genderList = []

    #
    # read the parish list 

    cellRow = 0
    cellCol = 2

    cell = ws.cell(row = cellRow,column = cellCol)
    while cell.value is not None:
        parishList.append(cell.value)
        cellCol = cellCol + 1
        cell = ws.cell(row = cellRow,column = cellCol)

    #
    # now get the grade list
    cellRow = 1
    cellCol = 0

    cell = ws.cell(row = cellRow,column = cellCol)
    while cell.value is not None:
        gradeList.append(cell.value)
        cellRow = cellRow + 1
        cell = ws.cell(row = cellRow,column = cellCol)

    #
    # finally get the gender list
    cellRow = 1
    cellCol = 1

    cell = ws.cell(row = cellRow,column = cellCol)
    while cell.value is not None:
        genderList.append(cell.value)
        cellRow = cellRow + 1
        cell = ws.cell(row = cellRow,column = cellCol)
    
    regDict = {}

    for index in range(len(gradeList)):
        tag = str(int(gradeList[index])) + genderList[index][0]
        regDict[tag] = []
        #numTeamsList = list(df.iloc[index,2:])
        numTeamsList = []
        cellRow = index + 1
        cellCol = 2
        for j in range(len(parishList)):
            cell = ws.cell(row = cellRow,column = cellCol)
            if cell.value is None:
                numTeamsList.append(0)
            else:
                numTeamsList.append(cell.value)
            cellCol = cellCol + 1
        for pIndex in range(0,len(parishList)):
            regDict[tag].append((parishList[pIndex],int(numTeamsList[pIndex])))

    return regDict

def readStandingsExcelFile(filename):
    #
    # open the file and set the active worksheet
    wb = load_workbook(filename)
    #
    # get list of active worksheets (should be 8B, 8G, 7B, 7G, etc.)

    listSheets = wb.get_sheet_names()

    dictAllGrades = {}


    for sheetName in listSheets:
        if len(listSheets) > 0:
            ws = wb.get_sheet_by_name(sheetName)

            #
            # initialize the standings dicitonary and set the grade

            dictStandings = {}
            currentGrade = int(sheetName[0])
            cellRow = 1
            cellCol = 0
            
            #
            # get the first row
            homeTeamCell = ws.cell(row = cellRow,column = cellCol)
            homeScoreCell = ws.cell(row = cellRow,column = cellCol+1)
            awayTeamCell = ws.cell(row = cellRow,column = cellCol+2)
            awayScoreCell = ws.cell(row = cellRow,column = cellCol+3)

            #
            # loop over all rows
            while homeTeamCell.value is not None:
                homeTeam = homeTeamCell.value.strip()
                homeScore = homeScoreCell.value
                awayTeam = awayTeamCell.value.strip()
                awayScore = awayScoreCell.value
                #
                # see if the keys exist 
                try:
                    if homeTeam in dictStandings.keys():
                        #
                        # team exists so just add info
                        dictStandings[homeTeam].addGame(awayTeam,'h')
                        dictStandings[homeTeam].setGameScore(homeScore, awayScore)
                    else:
                        #
                        # team does not exist, so create new object and add to dict
                        homeTeamObj = cmla.cmlaTeam()
                        homeTeamObj.setName(homeTeam)
                        homeTeamObj.setParish(homeTeam[0:3])
                        homeTeamObj.setGrade(currentGrade)
                        homeTeamObj.addGame(awayTeam,'h')
                        if homeTeam.lower() == "bye":
                            homeTeamObj.setGameScore(homeScore, awayScore, true)
                        else:
                            homeTeamObj.setGameScore(homeScore, awayScore)
                        dictStandings[homeTeam] = homeTeamObj

                    if awayTeam in dictStandings.keys():
                        #
                        # team exists so just add info
                        dictStandings[awayTeam].addGame(homeTeam,'a')
                        if homeTeam.lower() == "bye":
                            dictStandings[awayTeam].setGameScore(awayScore,homeScore, True)
                        else:
                            dictStandings[awayTeam].setGameScore(awayScore,homeScore)
                    else:
                        #
                        # team does not exist, so create new object and add to dict
                        awayTeamObj = cmla.cmlaTeam()
                        awayTeamObj.setName(awayTeam)
                        awayTeamObj.setParish(awayTeam[0:3])
                        awayTeamObj.setGrade(currentGrade)
                        awayTeamObj.addGame(homeTeam,'a')
                        if homeTeam.lower() == "bye":
                            awayTeamObj.setGameScore(awayScore, homeScore, True)
                        else:
                            awayTeamObj.setGameScore(awayScore, homeScore)
                        
                        dictStandings[awayTeam] = awayTeamObj
                except Exception:
                    print ('Exception ', Exception, homeTeam, awayTeam)

                cellRow = cellRow + 1
                #
                # get the next row of values
                homeTeamCell = ws.cell(row = cellRow,column = cellCol)
                homeScoreCell = ws.cell(row = cellRow,column = cellCol+1)
                awayTeamCell = ws.cell(row = cellRow,column = cellCol+2)
                awayScoreCell = ws.cell(row = cellRow,column = cellCol+3)

            dictAllGrades[sheetName] = dictStandings


    return dictAllGrades


def getTeamMatrix(scheduleTable):
#
#   make a table of teams
#   initialized to all zeroes

    teamMatrix = []

    for i in range(len(scheduleTable)):
        teamMatrix.append([])
        for j in range(len(scheduleTable)):
            teamMatrix[i].append(0)

    for pair in scheduleTable:
        (team1, team2) = pair
        teamMatrix[team1-1][team2-1] = 1
        teamMatrix[team2-1][team1-1] = 1
    return teamMatrix

def getTeamsPlayed(numTeams, scheduleTable):
    """
    This function will create a matrix of teams
    played from the scheudle table
    """
    teamsPlayed = []
    for i in range(numTeams):
        teamsPlayed.append([])

    for pair in (scheduleTable):
        (team1, team2) = pair
        teamsPlayed[team1].append(team2)
        teamsPlayed[team2].append(team1)

    return teamsPlayed

def getTeamsNotPlayed(numTeams, teamsPlayed):
    teamsNotPlayed = []
    allTeams = []
    for i in range(numTeams):
        teamsNotPlayed.append([])
        allTeams.append(i)

    for i in range(numTeams):
        teamsNotPlayed[i] = list(set(allTeams).difference(teamsPlayed[i]))
        teamsNotPlayed[i].remove(i)
        teamsNotPlayed[i].sort()

    return teamsNotPlayed
            
def makeRegistrationDict(filename):
    regDict = {}

    linenum = 0
    for line in open(filename,"r"):
        # print (line)
        tokens = line.split()
        linenum = linenum + 1
#        print (tokens)
        if len(tokens) > 1:
            if tokens[0] == "#":
                # comment line ignore
                print ('Comment ',line)
                pass
            else:
#                print (tokens)
                if len(tokens) == 3:
                    gradeGender = tokens[0]
                    parish = tokens[1]
                    numTeams = int(tokens[2])

                    if gradeGender in regDict.keys():
                        regDict[gradeGender].append((parish,numTeams))
                    else:
                        regDict[gradeGender] = []
                        regDict[gradeGender].append((parish,numTeams))


    return regDict


def makeTeamDict(teamFilename):
#
#   make a dictionary of all the teams
#   and initialize the object variables

    teamDict = {}

    linenum = 0
    for line in open(teamFilename,"r"):
        # print (line)
        tokens = line.split()
        linenum = linenum + 1
#        print (tokens)
        if len(tokens) > 1:
            if tokens[0] == "#":
                # comment line ignore
                print ('Comment ',line)
                pass
            else:
#                print (tokens)
                if len(tokens) == 2:
                    parish = tokens[0]
                    numTeams = int(tokens[1])
                    for i in range(numTeams):
                        newTeam = cmla.cmlaTeam()
                        teamName = parish +str(i+1)
                        newTeam.setName(teamName)
                        newTeam.setParish(parish)
                        teamDict[teamName] = newTeam
#    print ('In makeTeamDict',len(teamDict))
    if len(teamDict) % 2 == 1:
#
#       odd number of teams add BYE
        print ('Odd number of teams; Add BYE')
        newTeam = cmla.cmlaTeam()
        newTeam.setName("BYE")
        newTeam.setParish("BYE")
        teamDict["BYE"] = newTeam

    return teamDict

def makeGradeGenderTeamList(parishList):
#
#   make a dictionary of all the teams
#   and initialize the object variables

    teamDict = {}
    hasBye = False

    for pairs in parishList:
        (parish, numTeams) = pairs
        for i in range(numTeams):
            newTeam = cmla.cmlaTeam()
            teamName = parish +str(i+1)
            newTeam.setName(teamName)
            newTeam.setParish(parish)
            teamDict[teamName] = newTeam
#    print ('In makeTeamDict',len(teamDict))
    if len(teamDict) % 2 == 1:
#
#       odd number of teams add BYE
        print ('Odd number of teams; Add BYE')
        newTeam = cmla.cmlaTeam()
        newTeam.setName("BYE")
        newTeam.setParish("BYE")
        teamDict["BYE"] = newTeam
        hasBye = True

    return teamDict, hasBye


def makeSchedule(teamDict, teamsNotPlayed, teamsNotPlayedDict):

    #
    # get a list of teams and the number of teams
    teamList = list(teamDict.keys())
    numTeams = len(teamList)

    #
    # now make a random list of the teams to use for scheduling
    scheduleList = []
    emptySlots = []
    teamSlots = []
    for i in range(numTeams):
        scheduleList.append("empty")

    
    #
    # make a list of parishes

    parishDict = {}

    for team in teamList:
        if team[:3] in parishDict:
            parishDict[team[:3]] = parishDict[team[:3]] + 1
        else:
            parishDict[team[:3]] = 1
        
        
    trialNumber = 1
    currentIndex = 0
    makeList = True
    maxTeamsList = []
    
    while (len(teamList) > 0):
        try:
            if trialNumber > 5 and trialNumber < 10:
                #
                # choose bucket
                bucket = len(maxTeamsList) - 1
                parish = random.choice(maxTeamsList[bucket])
                maxTeamsList[bucket].remove(parish)
                if len(maxTeamsList[bucket]) == 0:
                    maxTeamsList.remove(maxTeamsList[bucket])
                team = parish + random.randint(1,bucket+1)


            elif trialNumber > 10:
                print ("Couldn't find solution --- Stopping  --- Schedule not Correct")
                makeList = True
                return scheduleList, 0
            else:
                #
                # pick a random team from teamList
                team = random.choice(teamList)

            #
            # get a list of empty slots
            emptySlots.clear()
            for i in range(len(scheduleList)):
                if scheduleList[i] == 'empty':
                    emptySlots.append(i)



            #
            # get number of teams submitted from parish

            numTeamsSubmitted = parishDict[team[:3]]

            #
            # if number of teams submitted = 1, pick 1st empty slot
            # if number of teams submitted = 2, use regular notPlayed list to find slots
            # if number of teams submitted > 2, use dictionary to try to find solution

            if numTeamsSubmitted == 1:
                #
                # add team to list in first available slot
                teamIndex = emptySlots[0]
                scheduleList[teamIndex] = team
                teamDict[team].setIndex(teamIndex)
                teamList.remove(team)
            elif numTeamsSubmitted == 2:
                parish = team[:3]
                if team == parish + "1":
                    team2 = parish + "2"
                else:
                    team2 = parish + "1"

                #
                # loop thru indicies in emptySlots to find a solution

                for index in emptySlots:
                    intersect = list(set(emptySlots).intersection(teamsNotPlayed[index]))

                    if len(intersect) >= 1:
                        #
                        # pull first slot for team1
                        teamIndex = index
                        scheduleList[teamIndex] = team
                        teamDict[team].setIndex(teamIndex)
                        teamList.remove(team)
                        teamIndex = random.choice(intersect)
                        scheduleList[teamIndex] = team2
                        teamDict[team2].setIndex(teamIndex)
                        teamList.remove(team2)
                        break
                    else:
                        raise Exception
            else:
                #
                # use the notPlayed list
                teamSlots.clear()

                ##
                ## have to map team indicies to internal indicies
                #for i in range(len(emptySlots)):
                #    teamSlots.append(emptySlots[i])
                #
                # more than 2 teams
                # create a local "notPlayed" from teh dictionary

                localNotPlayed = []
                teamIndexes = []
                for key in teamsNotPlayedDict:
                    tokens = key.split("T")
                    if len(tokens) == numTeamsSubmitted:
                        teamListing = []
                        for token in tokens:
                            if token != "":
                                teamListing.append(int(token,10))
                        for index in teamsNotPlayedDict[key]:
                            teamListing.append(index)
                        localNotPlayed.append(teamListing)

                    #
                    # if len of localnotplayed == 0, can't move on

                if len(localNotPlayed) > 0:
                    #
                    # now find a solution where the 
                    # teams overlap empty slots

                    solutionFound = False
                    for tlist in localNotPlayed:
                        solution = list(set(tlist).intersection(emptySlots))
                        

                        if len(solution) >= numTeamsSubmitted:
                            # possible solution!
                            # now check to see if 1st numTeamsSubmitted - 1 indexes (main ones for solution) are empty

                            mainIndexes = []
                            for i in range(numTeamsSubmitted-1):
                                mainIndexes.append(tlist[i])

                            if len(list(set(solution).intersection(mainIndexes))) == (numTeamsSubmitted - 1):
                                parish = team[:3]
                                pTeams = []
                                for i in range(1,numTeamsSubmitted+1):
                                    pTeams.append(parish + str(i))

                                for i in range(numTeamsSubmitted-1):
                                    pTeam = random.choice(pTeams)
                                    index = tlist[i]
                                    scheduleList[index] = pTeam
                                    teamDict[pTeam].setIndex(index)
                                    teamList.remove(pTeam)
                                    pTeams.remove(pTeam)
                                    solution.remove(index)


                                for pTeam in pTeams:
                                    index = random.choice(solution)
                                    scheduleList[index] = pTeam
                                    teamDict[pTeam].setIndex(index)
                                    teamList.remove(pTeam)
                                    pTeams.remove(pTeam)
                                    solution.remove(index)
                                solutionFound = True
                                break
                    if solutionFound == False:
                        raise Exception

                else:
                    raise Exception

        except Exception:
            trialNumber = trialNumber + 1

            #
            # reset the schedule list and team list
            scheduleList = []
            for i in range(numTeams):
                scheduleList.append("empty")
            teamList.clear()
            teamList = list(teamDict.keys())
            currentIndex = 0
            if trialNumber > 5 and trialNumber < 20:

                #
                # Haven't found a solution
                # try to choose the parishes with largest amount of 
                # teams submitted
                maxTeams = 0
                for key in parishDict:
                    if parishDict[key] > maxTeams:
                        maxTeams = parishDict[key]

                maxTeamsList = []
                for i in range (maxTeams):
                    maxTeamsList.append([])

                for key in parishDict:
                    maxTeamsList[parishDict[key] - 1].append(key)

            print ("    Raised exception --- trial # ",trialNumber)
            print ("    Resetting lists")



    return scheduleList,1

def updateCMLADict(cmlaDict, scheduleTable, teamList):
    
    index = 0
    for pair in scheduleTable:
        (away, home) = pair
        awayTeam = teamList[away]
        homeTeam = teamList[home]

        if awayTeam == "BYE":
            #
            # need to switch bye from being away to home team
            awayTeam = homeTeam
            homeTeam = "BYE"
            scheduleTable[index] = (home,away)

        cmlaDict[awayTeam].addGame(homeTeam,'a')
        cmlaDict[homeTeam].addGame(awayTeam,'h')
        index = index + 1
    return

def switchBookkeeping(switchHome, switchAway, cmlaDict, scheduleTable):
    switchHomeIndex = cmlaDict[switchHome].getIndex()
    switchAwayIndex = cmlaDict[switchAway].getIndex()
    
    switchIndex = 0
    #
    # find the pair in the scedule table
    for pair in scheduleTable:
        (away, home) = pair
        if (away == switchAwayIndex) and (home == switchHomeIndex):
            scheduleTable[switchIndex] = (home, away)
        else:
            switchIndex = switchIndex + 1
            
    cmlaDict[switchHome].flipHomeAway(switchAway)
    cmlaDict[switchAway].flipHomeAway(switchHome)
    return 

    

def loadBalanceSchedule(cmlaDict, scheduleTable, teamList):
    """
    This function will load balance the schedule so that each team 
    has 5 home and 5 away games
    """

    homeGamesList = []
    for i in range(10):
        homeGamesList.append([])

    for key in cmlaDict.keys():
        if key != "BYE":
            numHomeGames = cmlaDict[key].getNumHomeGames()
            homeGamesList[numHomeGames - 1].append(key)

    
    balanced = False
    currentIndex = 9
    while (currentIndex >= 5):
        if len(homeGamesList[currentIndex]) == 0:
            currentIndex = currentIndex - 1
        else:
#           
#           get the team to switch
            
            switchHomeTeam = random.choice(homeGamesList[currentIndex])
#
#           now get the list of teams played
            teamsPlayed = cmlaDict[switchHomeTeam].getHomeGamesList()
            switchFound = False
            searchIndex = 0
            while switchFound == False:
                switchBucket = 0 
                gotList = False
                switchbucket = 0
                while (gotList == False):
                    if len(homeGamesList[switchBucket]) == 0:
                        switchBucket = switchBucket + 1
                    else:
                        switchList = homeGamesList[switchBucket]
                        intersectList = list(set(teamsPlayed).intersection(switchList))
                        if len(intersectList) == 0:
                            switchBucket = switchBucket + 1
                        else:
                            gotList = True
                if len(intersectList) == 0:
                    #print ('No team found to switch home/away', currentIndex, searchIndex)
                    switchFound = True
                else:
                    #print ('Found these teams to switch with',intersectList)
                    #
                    # pick the first team and do the book keeping
                    
                    switchAwayTeam = random.choice(intersectList)
                    switchBookkeeping(switchHomeTeam, switchAwayTeam, cmlaDict, scheduleTable)
                    homeGamesList[currentIndex].remove(switchHomeTeam)
                    homeGamesList[currentIndex - 1].append(switchHomeTeam)
                    homeGamesList[switchBucket].remove(switchAwayTeam)
                    homeGamesList[switchBucket + 1].append(switchAwayTeam)
                    switchFound = True
#            currentIndex = currentIndex - 1
                    

    return scheduleTable

def writeExcelInterimFile(masterSchedule):

    excelFilename = tkinter.filedialog.asksaveasfilename()
    
    keyList = list(masterSchedule.keys())
    keyList.sort()
    #
    # create new workbook
    wb = Workbook()

    for key in keyList:
        #
        # create sheet in file for key
        ws = wb.create_sheet()
        ws.title = key
        (scheduleTable, scheduleList, hasBye, numTeams) = masterSchedule[key]
        sortedList = []
        for team in scheduleList:
            if team != "BYE":
                sortedList.append(team)
        sortedList.sort()
        if hasBye == True:
            sortedList.append("BYE")
        
        ws.cell(row=0,column=0).value =  key
        cellrow = 1
        cellcol = 0
        for team in sortedList:
            ws.cell(row=cellrow,column=cellcol).value = team
            cellrow = cellrow + 1

        teamSchedule = []

        cellrow = 0
        cellcol = 3
        ws.cell(row=cellrow,column=cellcol).value = 'Vis'
        ws.cell(row=cellrow,column=cellcol+1).value = 'Hm'
        cellrow = 1
        for game in scheduleTable:
            (away,home) = game
            ws.cell(row=cellrow,column=cellcol).value = scheduleList[away]
            ws.cell(row=cellrow,column=cellcol+1).value = scheduleList[home]
            cellrow = cellrow + 1

        ws.cell(row=0,column=10).value = 'NOTES:'
        ws.cell(row=1,column=10).value = '1) COPY A WORKSHEET WITH '+ str(numTeams) + ' TEAMS AND RENAME TO ' + key
        ws.cell(row=2,column=10).value = '2) COPY TEAM LIST [COL A] TO [COL M] ON SHEET ' + key
        ws.cell(row=3,column=10).value = "3) COPY SCHEDULE TABLE [COL'S D & E] TO [COL'S D & E] ON SHEET " + key

    wb.save(excelFilename)

    return

def buildNotPlayedDict(notPlayed, depth):
    """
    Function to build dictionary for multiple parish teams not played list 

    Format will be 

    key = T[#]T[#] *** T[#]
    value = list of possible teams not played

    ex,

    T1T5 = [3,7,9,10]

    means that team 1 and team 5 don't play teams 3,7,9 or team 10

    OR

    T1T5T9 = [16]

    means teams 1, 5, and 9 don't play team 16

    """

    notPlayedDict = {}
    numTeams = len(notPlayed)

    if depth >= 3:
        for i in range(numTeams):
            for j in range(len(notPlayed[i])):
                try:
                    teamIndex = notPlayed[i][j]
                    if teamIndex != i:
                        intersectList = list(set(notPlayed[i]).intersection(notPlayed[teamIndex]))
                        if len(intersectList) > 0:
                            #
                            # first check if key exists from symmetry
                            try:
                                keyexists = "T" + str(teamIndex) + "T" + str(i)
                                if keyexists in notPlayedDict.keys():
                                    pass
                                else:
                                    key = "T" + str(i) + "T" + str(teamIndex)
                                    notPlayedDict[key] = intersectList
                            except:
                                pass
                except:
                    pass

    tokens = []
    if depth >= 4:
        tempDict = {}
        for key in notPlayedDict:
            teamList = notPlayedDict[key]
            for j in range(len(teamList)):
                try:
                    tokens = str(key).split("T")
                    tokens = list(filter(None, tokens))
                    teamIndex = teamList[j]
                    intersectList = list(set(teamList).intersection(notPlayed[teamIndex]))
                    if len(intersectList) > 0:
                        if (teamIndex) < int(tokens[0],10):
                            newKey = "T" + str(teamIndex) + key
                        elif teamIndex < int(tokens[1],10):
                            newKey = "T" + str(tokens[0]) + "T" + str(teamIndex) + "T" + str(tokens[1])
                        else:
                            newKey = key + "T" + str(teamIndex)
                        if newKey in tempDict:
                            pass
                        else:
                            tempDict[newKey] = intersectList
                except:
                    pass
        for key in tempDict:
            notPlayedDict[key] = tempDict[key]

    return notPlayedDict

#
# This function sorts a dictionary based on winning percentage
def sortWinningPercentage(teamDict):
    """
    sortWinningPercentage will create list of keys and winning percentages
    from the information in the dictionary and then sort it according to 
    highest winning percentage. The function will return this sorted list

    dict    teamDict    dictionary of team data

    return value:   list    sortedList  sorted list of tuples continaing the team name and the winning percentage

    """
    sortList = []

    #
    # loop over all keys and create a tuple
    # of team name (key) and the winning
    # percentage

    for key in teamDict.keys():
        if key.lower() != 'bye':
            sortList.append((key,teamDict[key].getWinPercentage()))

    return sorted(sortList, key=lambda team: team[1],reverse=True)


#
# This function sorts a dictionary based on total point difference
def sortPointDifference(teamDict):
    """
    sortPointDifference will create list of keys and winning percentages
    from the information in the dictionary and then sort it according to 
    highest winning percentage. The function will return this sorted list

    dict    teamDict    dictionary of team data

    return value:   list    sortedList  sorted list of tuples continaing the team name and the point differential

    """
    sortList = []

    #
    # loop over all keys and create a tuple
    # of team name (key) and the winning
    # percentage

    for key in teamDict.keys():
        if key.lower() != 'bye':
            sortList.append((key,teamDict[key].getAdjustedPointDifferential()))

    return sorted(sortList, key=lambda team: team[1],reverse=True)

#
# calcSoS will calculate the SoS for all the teams in a dictionary

def calcSoS(teamDict):
    """
    calcSoS will calculate the strength of schedule for 
    all the teams in teamDict.  The SoS is calucated as 
    a weighted average of the following values:

        * Opponents winning percentage
        * SUM {[Opponents PF / [Total Games Played] } / [Max PF for all teams]
        * SUM {[Opponents PA / [Total Games Played] } / [Min PA for all teams]
    """

    #
    # First find the max PH and min PA 
    # for all teams

    maxPF = 0.0
    minPA = 1e4
    for team in teamDict.keys():
        if team.lower() != 'bye':
            if teamDict[team].getPFSeasonAvg() > maxPF: 
                maxPF = teamDict[team].getPFSeasonAvg()

            if teamDict[team].getPASeasonAvg() < minPA:
                minPA = teamDict[team].getPASeasonAvg()

    #
    # now loop over all teams and calcualte the SoS

    for team in teamDict.keys():
        if team.lower() != 'bye':
            oppList = teamDict[team].getOpponentList()
            sumWinPercent = 0.0
            sumPF = 0.0
            sumPA = 0.0
            for opponent in oppList:
                if opponent.lower() != 'bye':
                    sumWinPercent = sumWinPercent + teamDict[opponent].getWinPercentage()
                    sumPF = sumPF + (teamDict[opponent].getPFSeasonAvg() / maxPF)
                    try:
                        sumPA = sumPA + (minPA / teamDict[opponent].getPASeasonAvg())
                    except :
                        sumPA = 0.0
            totGames = teamDict[team].getTotalGamesPlayed()
            SoS = 60.0 * (sumWinPercent / totGames) + 20.0 * ((sumPF + sumPA)/totGames)
            teamDict[team].setSoS(SoS)

    return

#
# resolveTies will resolve ties based on CMLA rules

def resolveTies(divDict, tieList):
    """
    resolveTies will loop over the current standings 
    to find and resolve ties based on CMLA rules:

        Tournament Seeding & Regular Season Final Standings: Teams will be
        seeded for tournament play based on their regular season won-loss record. Teams who
        are tied at the end of the regular season will use the following tiebreaker:
        1. Head to Head record
        2. Head to Head point-differential (if more than 2 teams)
        3. Common Opponent(s) record
        4. Common Opponent(s) point-differential
        5. Season point differential (15 point or 12 point max. per game depending on grade)
        6. Coin Flip, or if more than two teams draw names out of a hat

    """
    #
    # set the intial tiebreaker number

    tieBreaker = 1

    resolvedList = []
    overallNotes = ""
    tieNotes = ""
    suceed = False
    notesList = []
    for i in range(len(tieList)):
        notesList.append("")

    while tieBreaker < 6:
        #
        # set current index for tieList
        cIndex = 0

        while cIndex < len(tieList):
            if isinstance(tieList[cIndex], list):
                if notesList[cIndex] == '':
                    notesList[cIndex] = "Seeds " + str(list(range(cIndex+1,cIndex+1 + len(tieList[cIndex])))) + ': \n'
                tieDict = createTiedDict(divDict, tieList[cIndex])
                if tieBreaker == 1:
                    #
                    # head to head
            
                    (succeed, resolvedList, tieNotes) = head2Head(tieDict)
                    
                elif tieBreaker == 2:
                    #
                    # head to head PD
                    (succeed, resolvedList, tieNotes) = head2HeadPDiff(tieDict)
                    
                elif tieBreaker == 3:
                    #
                    # common opponents record
                    (succeed, resolvedList, tieNotes) = commonOpponents(tieDict)
                elif tieBreaker == 4:
                    #
                    # common opponents PD
                    (succeed, resolvedList, tieNotes) = commonOpponentsPDiff(tieDict)
                elif tieBreaker == 5:
                    #
                    # overall PD
                    (succeed, resolvedList, tieNotes) = overallPDiff(tieDict)

                overallNotes = overallNotes + tieNotes
                notesList[cIndex] = notesList[cIndex] + " Tiebreaker " + str(tieBreaker) + ":" + tieNotes + " \n" 
                if succeed == True:
                    for item in resolvedList:
                        tieList[cIndex] = item
                        cIndex = cIndex + 1
                else:
                    cIndex = cIndex + len(tieList[cIndex])
            else:
                cIndex = cIndex + 1

        tieBreaker = tieBreaker + 1


    
    return (tieList, notesList)

#
# coallateList will find ties within a sorted list
# and place them into a sub list

def coallateList(sortedList):
    """
    coallateList will find ties within a sorted list
    and place them into a sub list in the apporpriate
    position within the list.  For example, if teams
    1,2 and 3 are tied the list will look like:
    [[1,2,3], 'empty', 'empty' ,...] 

        list    sortedList

        Return:     list    tiedList

    """

    newList = []

    #
    # initialize new list with "empty"

    numItems = len(sortedList)

    for index in range(numItems):
        newList.append('empty')

    tIndex = 0
    while tIndex < numItems:
        tVal  = sortedList[tIndex][1]
        tieList = []
        tieList.append(sortedList[tIndex])
        cIndex = tIndex + 1
        #
        # find the tied values
        while (cIndex < numItems) and (math.fabs(sortedList[cIndex][1] - tVal) < 0.001):
            tieList.append(sortedList[cIndex])
            cIndex = cIndex + 1

        #
        # done with ties
        if len(tieList) == 1:
            newList[tIndex] = sortedList[tIndex]
        else:
            newList[tIndex] = tieList
        tIndex = cIndex

    #
    # done looping over list
    return newList


#
# head to head tie resolution
def head2Head(teamDict):
    """
    head2Head will try to resolve the teams that 
    are tied (listed in tiedTeams list) by comparing
    head to head records.  The function will
    return a tuple containing a boolean value if 
    successful, the sorted list, and any notes about 
    the resolution

        dict    teamDict    - dictionary of tied teams

        Return value:
            tuple   (True | False, list: sorted teams, string: Notes)

    """

    succeed = True
    played = []
    tiedTeams = list(teamDict.keys())
    numTied = len(tiedTeams)
    notes = ""
    tiedSort = []
    h2HCoallate = []
    
    for team in tiedTeams:
        listOpp = teamDict[team].getOpponentList()
        played.clear()

        played = set(listOpp).intersection(tiedTeams)

        if len(played) == numTied - 1:
            # current team played h2h with other tied teams
            pass
            #newTeam = cmla.cmlaTeam()
            #newTeam.setName(team)
            #for addTeam in played:
            #    (index, location, (teamScore, oppScore)) = teamDict[team].getOpponentInfo(addTeam)
            #    if index > 0:
            #        newTeam.addGame(addTeam, location)
            #        newTeam.setGameScore(teamScore, oppScore)
            #tiedDict[team] = newTeam
        else:
            succeed = False
            notes = "H2H: " + ', '.join(tiedTeams) + " did not play"
            break

    if succeed:
        #
        # sort list based on winning percentage
        tiedDict = createCommonDict(teamDict, tiedTeams)

        tiedSort = sortWinningPercentage(tiedDict)
        notes = " H2H: " + str(tiedSort)
        
        h2HCoallate = coallateList(tiedSort)


    return (succeed, h2HCoallate, notes)

#
# head to head point differential tie resolution
def head2HeadPDiff(teamDict):
    """
    head2Head will try to resolve the teams that 
    are tied (listed in tiedTeams list) by comparing
    head to head point differentials.  The function will
    return a tuple containing a boolean value if 
    successful, the sorted list, and any notes about 
    the resolution

        dict    teamDict    - dictionary of all teams results
        list    tiedTeams   - list of teams that are tied

        Return value:
            tuple   (True | False, list: sorted teams, string: Notes)

    """

    succeed = True
    played = []
    tiedTeams = list(teamDict.keys())
    numTied = len(tiedTeams)
    notes = ""
    tiedSort = []
    h2HCoallate = []

    for team in tiedTeams:
        listOpp = teamDict[team].getOpponentList()
        played.clear()

        played = set(listOpp).intersection(tiedTeams)

        if len(played) == numTied - 1:
            pass
            #tiedDict = createCommonDict(teamDict, played)
            #newTeam = cmla.cmlaTeam()
            #newTeam.setName(team)
            #for addTeam in played:
            #    (index, location, (teamScore, oppScore)) = teamDict[team].getOpponentInfo(addTeam)
            #    if index > 0:
            #        newTeam.addGame(addTeam, location)
            #        newTeam.setGameScore(teamScore, oppScore)
            #tiedDict[team] = newTeam
        else:
            succeed = False
            notes = "H2H PD: " + ', '.join(tiedTeams) + " did not play H2H"
            break

    if succeed:
        #
        # sort list based on winning percentage
        tiedDict = createCommonDict(teamDict, tiedTeams)
        tiedSort = sortPointDifference(tiedDict)
        notes = " H2HPD: " + str(tiedSort)
        
        h2HCoallate = coallateList(tiedSort)


    return (succeed, h2HCoallate, notes)

#
# createCommonDict will create a new season dictionary 
# based on the parameter commonList
 
def createCommonDict(tiedDict, commonList):
    """
    createCommonDict will create a new 'season' dictionary
    that includes only the opponents listed in commonList.
    this function will be used to resolve ties for 
    season records and point differentials against common opponents

        dict    tiedDict    contains the tied teams and their season info
        list    commonList  contains the common opponents for the tied teams
        
        Return value:   dict    commonDict  dictionary of common opponets 
    """

    commonDict = {}
    for key in tiedDict.keys():
        newTeam = cmla.cmlaTeam()
        newTeam.setName(key)
        newTeam.setGrade(int(key[3]))

        for team in commonList:
            (index, location, (teamScore, oppScore)) = tiedDict[key].getOpponentInfo(team)
            if index >= 0:
                newTeam.addGame(team, location)
                newTeam.setGameScore(teamScore, oppScore)
                
        commonDict[key] = newTeam 

    return commonDict

#
# commonOpponents will resolve ties based on winning %
# against common opponents

def commonOpponents(tiedDict):
    """
    commonOpponent will resolve a tie based on the winning percentage
    of tied teams versus common opponents

        dict    tiedDict    dictionary of tied teams
        
        Return:
        Return value:
            tuple   (True | False, list: sorted teams, string: Notes)            


    """
    succeed = True
    played = []
    tiedTeams = list(tiedDict.keys())
    numTied = len(tiedTeams)
    notes = ""
    tiedSort = []
    h2HCoallate = []
    played = tiedDict[tiedTeams[0]].getOpponentList()
        
    for index in range(1,len(tiedTeams)):
        listOpp = tiedDict[tiedTeams[index]].getOpponentList()
        played = set(listOpp).intersection(played)

        if len(played) == 0:
            succeed = False
            notes = ', '.join(tiedTeams) + " have no common opponents"
            break

    if succeed:
        #
        # sort list based on winning percentage
        tiedCommonDict = createCommonDict(tiedDict, played)
        tiedSort = sortWinningPercentage(tiedCommonDict)
        notes = " Common Opp: " + str(tiedSort)
        h2HCoallate = coallateList(tiedSort)
    
    return (succeed, h2HCoallate, notes)


#
# commonOpponentsPDiff will resolve ties based on point differential
# against common opponents

def commonOpponentsPDiff(tiedDict):
    """
    commonOpponent will resolve a tie based on the point differential
    of tied teams versus common opponents

        dict    tiedDict    dictionary of tied teams
        
        Return:
        Return value:
            tuple   (True | False, list: sorted teams, string: Notes)            


    """
    succeed = True
    played = []
    tiedTeams = list(tiedDict.keys())
    numTied = len(tiedTeams)
    notes = ""
    tiedSort = []
    h2HCoallate = []
    played = tiedDict[tiedTeams[0]].getOpponentList()
        
    for index in range(1,len(tiedTeams)):
        listOpp = tiedDict[tiedTeams[index]].getOpponentList()
        played = set(listOpp).intersection(played)

        if len(played) == 0:
            succeed = False
            notes = ', '.join(tiedTeams) + " have no common opponents"
            break

    if succeed:
        #
        # sort list based on winning percentage
        tiedCommonDict = createCommonDict(tiedDict, played)
        tiedSort = sortPointDifference(tiedCommonDict)
        notes = " Common Opp PD: " + str(tiedSort)
        h2HCoallate = coallateList(tiedSort)
    return (succeed, h2HCoallate, notes)

#
# overallPDiff will resolve ties based on point differential
# for the whole season

def overallPDiff(tiedDict):
    """
    overallPDiff will resolve a tie based on the point differential
    over the whole season

        dict    tiedDict    dictionary of tied teams
        
        Return:
        Return value:
            tuple   (True | False, list: sorted teams, string: Notes)            


    """
    succeed = True
    played = []

    tiedTeams = list(tiedDict.keys())
    notes = ""
    tiedSort = []
    h2HCoallate = []
        
    #
    # sort list based on overall point difference

    tiedSort = sortPointDifference(tiedDict)
    h2HCoallate = coallateList(tiedSort)
    notes = " Overall PD: " + str(tiedSort)
    
    for item in h2HCoallate:
        if isinstance(item, list):
            succeed = False
            notes = ' Overall PD:  '.join(tiedTeams) + " still tied, need coin toss"
            break

    return (succeed, h2HCoallate, notes)

#
# --------------------------------------------------------------------------
#

# 
# createTiedDict will make a new dictionary 
# for tied teams supplied in a list 

def createTiedDict(divDict, tiedList):
    """
    createTiedDict will make a new dictionary
    containing only the teams that are listed 
    in the parameter tiedList.  This dictionary is 
    then used in the tie breaker functions 

    dict    divDict     entire division dictionary
    list    tiedList    list of tied team tuples (team name, value)

    Return value:
        dict    tiedDict    dictionary containing only tied teams

    """
    tiedDict = {}
    for teamTuple in tiedList:
        tiedDict[teamTuple[0]] = divDict[teamTuple[0]]

    return tiedDict



def main():
    """
    main is the main function for the cmla program.  it will take several 
    command line arguments to either build a season schedule or compute the
    season ending standings for a division

    arguments:
        -h              :   help on running the code
        -c              :   compute season ending standings
        -s              :   compute season schedule
        -f [filename]   :   standings filename
        -n [# teams]    :   number of teams
        -t [filename]   :   team list filename
        -r [filename]   :   registration filename
        -y [year]       :   year 

    """
    teamListFilename = ""
    registrationFilename = ""
    standingsFilename = ""
    scheduleGames = False
    computeStandings = False
    standingsYear = 0
    try:
        opts, args = getopt.getopt(sys.argv[1:], "chsf:n:t:r:y:",["help","filename="])
    except getopt.error as msg:
        print (msg)
        print ("for help use --help")
        sys.exit(2)

    for o, arg in opts:
#        print (o, arg)
        if o == "-h":
            print ("python cmla.py")
        if o == "-f":
            standingsFile = arg
        if o == "-n":
            numTeams = int(arg,10)
        if o == "-t":
            teamListFilename = arg
        if o == "-r":
            registrationFilename = arg
        if o == "-s":
            scheduleGames = True
        if o == "-c":
            computeStandings = True
        if o == "-y":
            standingsYear = int(arg,10)


    if scheduleGames == True:
        if registrationFilename == "":
            registrationFilename = tkinter.filedialog.askopenfilename()

        tokens = registrationFilename.split('.')
        numTokens = len(tokens)
        extension = tokens[numTokens-1]

        if extension in ('xls','xlsx'):
            regDict = readRegistrationExcelFile(registrationFilename)
        else:
            regDict = makeRegistrationDict(registrationFilename)
    #       print ("Team list is ", teamListFilename)

        #
        # Now loop over all grade/genders to make the schedule
        rKeys = list(regDict.keys())
        rKeys.sort()
        masterSchedule = {}
        notPlayedDict = {}
        for rKey in rKeys:
            notPlayedDict.clear()
            if rKey[0] == "8":
                print (" Starting 8's")
            parishList = regDict[rKey]
            cmlaTeamDict, hasBye = makeGradeGenderTeamList(parishList)
            numTeams = len(cmlaTeamDict)
            scheduleTable = getScheduleTable(len(cmlaTeamDict))
            teamsPlayed = getTeamsPlayed(numTeams, scheduleTable)
            teamsNotPlayed = getTeamsNotPlayed(numTeams, teamsPlayed)
            notPlayedDict = buildNotPlayedDict(teamsNotPlayed, 4)
            (scheduleList, completed) = makeSchedule(cmlaTeamDict, teamsNotPlayed, notPlayedDict)
            if completed == 1:
                updateCMLADict(cmlaTeamDict, scheduleTable, scheduleList)
                scheduleTable = loadBalanceSchedule(cmlaTeamDict, scheduleTable, scheduleList)
                print('##########################')
                print('      ',rKey,' Schedule   ')
                print('##########################')
                tKeys = list(cmlaTeamDict.keys())
                tKeys.sort()
                for key in tKeys:
                    print ('Team ', key, ' has ',cmlaTeamDict[key].getNumHomeGames(),'home games')
                    print ('     ',cmlaTeamDict[key].getHomeGamesList())
                    print ('  and ',cmlaTeamDict[key].getNumAwayGames(),'away games')
                    print ('     ',cmlaTeamDict[key].getAwayGamesList())

                print()
                print()
                masterSchedule[rKey] = (scheduleTable, scheduleList, hasBye, len(scheduleList))
        writeExcelInterimFile(masterSchedule)
    elif computeStandings == True:
        print ("Computing Standings")
        if standingsFilename == "":
            standingsFilename = tkinter.filedialog.askopenfilename()
        tokens = standingsFilename.split('.')
        numTokens = len(tokens)
        extension = tokens[numTokens-1]
        standingsYear = tokens[0]
        if extension in ('xls','xlsx'):
            #
            # read the season ending data into the standings
            # dictionary.  Each division is a key, i.e., 8B, 8G, etc..

            standingsDict = readStandingsExcelFile(standingsFilename)
            print ('Completed reading standings file')

            #
            # loop over all divisions and:
            #   1) calcuate the Stength of schedule
            #   2) Sort by winning percentages
            #   3) resolve ties based on CMLA rules
            divDict = {}
            divNotes = {}
            divStandings = {}
            for key in standingsDict.keys():
                divDict.clear()
                divDict = dict.copy(standingsDict[key])
                calcSoS(divDict)
                teamSorted = sortWinningPercentage(divDict)
                tieSorted = coallateList(teamSorted)
                print ('Completed Phase 1: sorted ',key, ' by winning percentage ')

                #
                # Now loop thru the tiedSorted list to find any tied teams 
                # and pass them to the resolveTies function

                teamIndex = 1
                tiedDict = {}

                #for item in tieSorted:
                #    if isinstance(item, list):
                #        #
                #        # found a tie, make a tiedDict and then call
                #        # resolveTies
                #        tiedDict = createTiedDict(divDict, item)
                (divStandings[key], divNotes[key]) = resolveTies(divDict, tieSorted)

            #
            # now write the data back out to text files

            for key in standingsDict.keys():
                f = open(key + ".txt",'w+')
                index = 1
                f.write("Seed\tTeam\tWins\tLosses\tTies\tWP\tPD\tSOS\n")
                f.write("----\t----\t----\t------\t----\t--\t--\t---\n")
                for (team,value) in divStandings[key]:
                    f.write(str(index) + "\t" + team + "\t" +   str(standingsDict[key][team].getWins()) +   "\t"  +  str(standingsDict[key][team].getLosses()) +   "\t" +  str(standingsDict[key][team].getTies()) +  "\t")
                    f.write('{:.2%}'.format(standingsDict[key][team].getWinPercentage()) +   "\t"  +  str(standingsDict[key][team].getAdjustedPointDifferential()) +   "\t" +  '{:.2%}'.format(standingsDict[key][team].getSoS()/100) +  "\n")
                    index = index + 1

                f.write("\nNOTES\n")
                f.write("-----\n")
                index = 1
                for note in divNotes[key]:
                    if note != '':
                        f.write(str(index) + ". " + note + '\n\n')
                        index = index + 1
                f.close()



                        
        else:
            pass
    else:
        print ('No task specified -- stopping')

if __name__ == "__main__":
    main()
