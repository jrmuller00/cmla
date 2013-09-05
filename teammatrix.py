import sys
import getopt
import cmlaTeam as cmla
import random
import tkinter
from tkinter import messagebox
from tkinter import filedialog
from pandas.io.parsers import ExcelFile
import pandas as pd
from openpyxl import Workbook


def getScheduleTable(numTeams):
#
#   make a table of teams
#   initialized to all zeroes
#   Note that team numbers are indexed
#   to zero 

    Filename = str(numTeams) + "teams"
    try:
        if len(scheduleTable) > 0:
            scheduleTable.clear()
    except Exception:
        scheduleTable = []

    linenum = 0
    for line in open(Filename,"r"):
        # print (line)
        tokens = line.split()
        linenum = linenum + 1
#        print (tokens)
        if len(tokens) > 1:
            if tokens[0] == "#":
                # comment line ignore
                print ('Comment ',line)
                pass
            else:
#                print (tokens)
                if len(tokens) == 3:
                    team1 = int(tokens[0]) - 1
                    team2 = int(tokens[2]) - 1
                    scheduleTable.append((team1, team2))
    return scheduleTable

def readExcelFile(filename):
    xd = pd.ExcelFile(filename)
    df = xd.parse('Sheet1')
    df.fillna(0,inplace=True)
    parishList = list(df.columns[2:])
    gradeList = list(df.iloc[0:,0])
    genderList = list(df.iloc[0:,1])
    regDict = {}
    for index in range(len(gradeList)):
        tag = str(int(gradeList[index])) + genderList[index][0]
        regDict[tag] = []
        numTeamsList = list(df.iloc[index,2:])
        for pIndex in range(0,len(parishList)):
            regDict[tag].append((parishList[pIndex],int(numTeamsList[pIndex])))

    return regDict


def getTeamMatrix(scheduleTable):
#
#   make a table of teams
#   initialized to all zeroes

    teamMatrix = []

    for i in range(len(scheduleTable)):
        teamMatrix.append([])
        for j in range(len(scheduleTable)):
            teamMatrix[i].append(0)

    for pair in scheduleTable:
        (team1, team2) = pair
        teamMatrix[team1-1][team2-1] = 1
        teamMatrix[team2-1][team1-1] = 1
    return teamMatrix

def getTeamsPlayed(numTeams, scheduleTable):
    """
    This function will create a matrix of teams
    played from the scheudle table
    """
    teamsPlayed = []
    for i in range(numTeams):
        teamsPlayed.append([])

    for pair in (scheduleTable):
        (team1, team2) = pair
        teamsPlayed[team1].append(team2)
        teamsPlayed[team2].append(team1)

    return teamsPlayed

def getTeamsNotPlayed(numTeams, teamsPlayed):
    teamsNotPlayed = []
    allTeams = []
    for i in range(numTeams):
        teamsNotPlayed.append([])
        allTeams.append(i)

    for i in range(numTeams):
        teamsNotPlayed[i] = list(set(allTeams).difference(teamsPlayed[i]))
        teamsNotPlayed[i].remove(i)
        teamsNotPlayed[i].sort()

    return teamsNotPlayed
            
def makeRegistrationDict(filename):
    regDict = {}

    linenum = 0
    for line in open(filename,"r"):
        # print (line)
        tokens = line.split()
        linenum = linenum + 1
#        print (tokens)
        if len(tokens) > 1:
            if tokens[0] == "#":
                # comment line ignore
                print ('Comment ',line)
                pass
            else:
#                print (tokens)
                if len(tokens) == 3:
                    gradeGender = tokens[0]
                    parish = tokens[1]
                    numTeams = int(tokens[2])

                    if gradeGender in regDict.keys():
                        regDict[gradeGender].append((parish,numTeams))
                    else:
                        regDict[gradeGender] = []
                        regDict[gradeGender].append((parish,numTeams))


    return regDict


def makeTeamDict(teamFilename):
#
#   make a dictionary of all the teams
#   and initialize the object variables

    teamDict = {}

    linenum = 0
    for line in open(teamFilename,"r"):
        # print (line)
        tokens = line.split()
        linenum = linenum + 1
#        print (tokens)
        if len(tokens) > 1:
            if tokens[0] == "#":
                # comment line ignore
                print ('Comment ',line)
                pass
            else:
#                print (tokens)
                if len(tokens) == 2:
                    parish = tokens[0]
                    numTeams = int(tokens[1])
                    for i in range(numTeams):
                        newTeam = cmla.cmlaTeam()
                        teamName = parish +str(i+1)
                        newTeam.setName(teamName)
                        newTeam.setParish(parish)
                        teamDict[teamName] = newTeam
#    print ('In makeTeamDict',len(teamDict))
    if len(teamDict) % 2 == 1:
#
#       odd number of teams add BYE
        print ('Odd number of teams; Add BYE')
        newTeam = cmla.cmlaTeam()
        newTeam.setName("BYE")
        newTeam.setParish("BYE")
        teamDict["BYE"] = newTeam

    return teamDict

def makeGradeGenderTeamList(parishList):
#
#   make a dictionary of all the teams
#   and initialize the object variables

    teamDict = {}
    hasBye = False

    for pairs in parishList:
        (parish, numTeams) = pairs
        for i in range(numTeams):
            newTeam = cmla.cmlaTeam()
            teamName = parish +str(i+1)
            newTeam.setName(teamName)
            newTeam.setParish(parish)
            teamDict[teamName] = newTeam
#    print ('In makeTeamDict',len(teamDict))
    if len(teamDict) % 2 == 1:
#
#       odd number of teams add BYE
        print ('Odd number of teams; Add BYE')
        newTeam = cmla.cmlaTeam()
        newTeam.setName("BYE")
        newTeam.setParish("BYE")
        teamDict["BYE"] = newTeam
        hasBye = True

    return teamDict, hasBye

#def makeSchedule(teamDict, teamsNotPlayed, teamsNotPlayedDict):

#    #
#    # get a list of teams and the number of teams
#    teamList = list(teamDict.keys())
#    numTeams = len(teamList)

#    #
#    # now make a random list of the teams to use for scheduling
#    scheduleList = []
#    for i in range(numTeams):
#        scheduleList.append("empty")

    
#    #
#    # make a list of parishes

#    parishDict = {}

#    for team in teamList:
#        if team[:3] in parishDict:
#            parishDict[team[:3]] = parishDict[team[:3]] + 1
#        else:
#            parishDict[team[:3]] = 1
        
        
#    trialNumber = 1
#    currentIndex = 0
#    makeList = True
    
#    while (makeList == True):
#        try:
#            if trialNumber > 15:
#                #
#                # reset the schedule list and team list
#                scheduleList = []
#                for i in range(numTeams):
#                    scheduleList.append("empty")
#                teamList.clear()
#                teamList = list(teamDict.keys())
#                currentIndex = 0

#                #
#                # sort teams into buckets according to how many each parish submitted
#                maxParish = teamList[0][:3]
#                maxTeams = parishDict[maxParish]
#                for teamCount in teamList:
#                    parish = teamCount[:3]
#                    if parishDict[parish] > maxTeams:
#                        maxParish = parish
#                        maxTeams = parishDict[parish]


#                #
#                # now with the maxTeams parish add these teams to a list and choose one

#                maxTeamsList = []
#                for teams in teamList:
#                    if teams[:3] == maxParish:
#                        maxTeamsList.append(teams)

#                team = random.choice(maxTeamsList)
#            elif trialNumber > 25:
#                print ("Couldn't find solution --- Stopping  --- Schedule not Correct")
#                makeList = True
#                return
#            else:
#                team = random.choice(teamList)
#            # 
#            # add the team to the schedule list
#            while scheduleList[currentIndex] != "empty":
#                currentIndex = currentIndex + 1
#            scheduleList[currentIndex] = team
#            teamDict[team].setListIndex(currentIndex)
#            teamList.remove(team)
#            #
#            # now need to find any other teams from that parish
#            # and add them to the schedule list

#            noPlayIndicies = []
#            noPlayIndicies.append(currentIndex)

#            sameParish = []
#            for otherTeam in teamList:
#                if team[:3] == otherTeam[:3]:
#    #                print ('Found same parish ',team, otherTeam)
#                    #
#                    # need to pull them from the team list
#                    # and add them to the schedule list in a slot that 
#                    # does not play "team"
#                    sameParish.append(otherTeam)

#            notPlayed = teamsNotPlayed[currentIndex]
#            sameParishIndex = []
#            for j in range(len(sameParish)):
#                undoTeam = False
#                try:
#                    if len(notPlayed) == 0: 
#                        print ("Error can't find solution to multiple parish teams")
#                    else:
#                        emptySlot = False
#                        k = 0
#                        intersectNotPlayed = []
#                        while emptySlot == False:
#                            spi = notPlayed[k] - 1
#                            if scheduleList[spi] == 'empty':
#                                sameParishIndex.append(spi)
#                                intersectNotPlayed = set(notPlayed).intersection(teamsNotPlayed[spi])
#                                notPlayed = list(intersectNotPlayed)
#                                emptySlot = True
#                            else:
#                                k = k + 1
#                except Exception:
#                    undoTeam = True

#            if (len(sameParish) == len(sameParishIndex)):
#                for j in range(len(sameParish)):
#                    otherTeam = sameParish[j]
#                    spi = random.choice(sameParishIndex)
#                    scheduleList[spi] = otherTeam
#                    teamDict[otherTeam].setListIndex(spi)
#                    teamList.remove(otherTeam)
#                    sameParishIndex.remove(spi)
#            else:
#                undoTeam = True
#                raise Exception("Error not enough indicies for same parish teams")
        
#            if len(teamList) == 0:
#                makeList = False

#        except Exception:
#            trialNumber = trialNumber + 1
#            #
#            if undoTeam == True:
#                scheduleList[currentIndex] = 'empty'
#                teamDict[team].setListIndex(-1)
#                teamList.append(team)
#            print ("    Raised exception --- trial # ",trialNumber)

#    return scheduleList


def makeSchedule(teamDict, teamsNotPlayed, teamsNotPlayedDict):

    #
    # get a list of teams and the number of teams
    teamList = list(teamDict.keys())
    numTeams = len(teamList)

    #
    # now make a random list of the teams to use for scheduling
    scheduleList = []
    emptySlots = []
    teamSlots = []
    for i in range(numTeams):
        scheduleList.append("empty")

    
    #
    # make a list of parishes

    parishDict = {}

    for team in teamList:
        if team[:3] in parishDict:
            parishDict[team[:3]] = parishDict[team[:3]] + 1
        else:
            parishDict[team[:3]] = 1
        
        
    trialNumber = 1
    currentIndex = 0
    makeList = True
    maxTeamsList = []
    
    while (len(teamList) > 0):
        try:
            if trialNumber > 5 and trialNumber < 10:
                #
                # choose bucket
                bucket = len(maxTeamsList) - 1
                parish = random.choice(maxTeamsList[bucket])
                maxTeamsList[bucket].remove(parish)
                if len(maxTeamsList[bucket]) == 0:
                    maxTeamsList.remove(maxTeamsList[bucket])
                team = parish + random.randint(1,bucket+1)


            elif trialNumber > 10:
                print ("Couldn't find solution --- Stopping  --- Schedule not Correct")
                makeList = True
                return
            else:
                #
                # pick a random team from teamList
                team = random.choice(teamList)

            #
            # get a list of empty slots
            emptySlots.clear()
            for i in range(len(scheduleList)):
                if scheduleList[i] == 'empty':
                    emptySlots.append(i)



            #
            # get number of teams submitted from parish

            numTeamsSubmitted = parishDict[team[:3]]

            #
            # if number of teams submitted = 1, pick 1st empty slot
            # if number of teams submitted = 2, use regular notPlayed list to find slots
            # if number of teams submitted > 2, use dictionary to try to find solution

            if numTeamsSubmitted == 1:
                #
                # add team to list in first available slot
                teamIndex = emptySlots[0]
                scheduleList[teamIndex] = team
                teamDict[team].setListIndex(teamIndex)
                teamList.remove(team)
            elif numTeamsSubmitted == 2:
                parish = team[:3]
                if team == parish + "1":
                    team2 = parish + "2"
                else:
                    team2 = parish + "1"

                #
                # loop thru indicies in emptySlots to find a solution

                for index in emptySlots:
                    intersect = list(set(emptySlots).intersection(teamsNotPlayed[index]))

                    if len(intersect) >= 1:
                        #
                        # pull first slot for team1
                        teamIndex = index
                        scheduleList[teamIndex] = team
                        teamDict[team].setListIndex(teamIndex)
                        teamList.remove(team)
                        teamIndex = random.choice(intersect)
                        scheduleList[teamIndex] = team2
                        teamDict[team2].setListIndex(teamIndex)
                        teamList.remove(team2)
                        break
                    else:
                        raise Exception
            else:
                #
                # use the notPlayed list
                teamSlots.clear()

                ##
                ## have to map team indicies to internal indicies
                #for i in range(len(emptySlots)):
                #    teamSlots.append(emptySlots[i])
                #
                # more than 2 teams
                # create a local "notPlayed" from teh dictionary

                localNotPlayed = []
                teamIndexes = []
                for key in teamsNotPlayedDict:
                    tokens = key.split("T")
                    if len(tokens) == numTeamsSubmitted:
                        teamListing = []
                        for token in tokens:
                            if token != "":
                                teamListing.append(int(token,10))
                        for index in teamsNotPlayedDict[key]:
                            teamListing.append(index)
                        localNotPlayed.append(teamListing)

                    #
                    # if len of localnotplayed == 0, can't move on

                if len(localNotPlayed) > 0:
                    #
                    # now find a solution where the 
                    # teams overlap empty slots

                    solutionFound = False
                    for tlist in localNotPlayed:
                        solution = list(set(tlist).intersection(emptySlots))
                        

                        if len(solution) >= numTeamsSubmitted:
                            # possible solution!
                            # now check to see if 1st numTeamsSubmitted - 1 indexes (main ones for solution) are empty

                            mainIndexes = []
                            for i in range(numTeamsSubmitted-1):
                                mainIndexes.append(tlist[i])

                            if len(list(set(solution).intersection(mainIndexes))) == (numTeamsSubmitted - 1):
                                parish = team[:3]
                                pTeams = []
                                for i in range(1,numTeamsSubmitted+1):
                                    pTeams.append(parish + str(i))

                                for i in range(numTeamsSubmitted-1):
                                    pTeam = random.choice(pTeams)
                                    index = tlist[i]
                                    scheduleList[index] = pTeam
                                    teamDict[pTeam].setListIndex(index)
                                    teamList.remove(pTeam)
                                    pTeams.remove(pTeam)
                                    solution.remove(index)


                                for pTeam in pTeams:
                                    index = random.choice(solution)
                                    scheduleList[index] = pTeam
                                    teamDict[pTeam].setListIndex(index)
                                    teamList.remove(pTeam)
                                    pTeams.remove(pTeam)
                                    solution.remove(index)
                                solutionFound = True
                                break
                    if solutionFound == False:
                        raise Exception

                else:
                    raise Exception

        except Exception:
            trialNumber = trialNumber + 1

            #
            # reset the schedule list and team list
            scheduleList = []
            for i in range(numTeams):
                scheduleList.append("empty")
            teamList.clear()
            teamList = list(teamDict.keys())
            currentIndex = 0
            if trialNumber > 5 and trialNumber < 20:

                #
                # Haven't found a solution
                # try to choose the parishes with largest amount of 
                # teams submitted
                maxTeams = 0
                for key in parishDict:
                    if parishDict[key] > maxTeams:
                        maxTeams = parishDict[key]

                maxTeamsList = []
                for i in range (maxTeams):
                    maxTeamsList.append([])

                for key in parishDict:
                    maxTeamsList[parishDict[key] - 1].append(key)

            print ("    Raised exception --- trial # ",trialNumber)
            print ("    Resetting lists")



    return scheduleList

def updateCMLADict(cmlaDict, scheduleTable, teamList):
    
    index = 0
    for pair in scheduleTable:
        (away, home) = pair
        awayTeam = teamList[away]
        homeTeam = teamList[home]

        if awayTeam == "BYE":
            #
            # need to switch bye from being away to home team
            awayTeam = homeTeam
            homeTeam = "BYE"
            scheduleTable[index] = (home,away)

        cmlaDict[awayTeam].addGame(homeTeam,'a')
        cmlaDict[homeTeam].addGame(awayTeam,'h')
        index = index + 1
    return

def switchBookkeeping(switchHome, switchAway, cmlaDict, scheduleTable):
    switchHomeIndex = cmlaDict[switchHome].getListIndex()
    switchAwayIndex = cmlaDict[switchAway].getListIndex()
    
    switchIndex = 0
    #
    # find the pair in the scedule table
    for pair in scheduleTable:
        (away, home) = pair
        if (away == switchAwayIndex) and (home == switchHomeIndex):
            scheduleTable[switchIndex] = (home, away)
        else:
            switchIndex = switchIndex + 1
            
    cmlaDict[switchHome].flipHomeAway(switchAway)
    cmlaDict[switchAway].flipHomeAway(switchHome)
    return 

    

def loadBalanceSchedule(cmlaDict, scheduleTable, teamList):
    """
    This function will load balance the schedule so that each team 
    has 5 home and 5 away games
    """

    homeGamesList = []
    for i in range(10):
        homeGamesList.append([])

    for key in cmlaDict.keys():
        if key != "BYE":
            numHomeGames = cmlaDict[key].getNumHomeGames()
            homeGamesList[numHomeGames - 1].append(key)

    
    balanced = False
    currentIndex = 9
    while (currentIndex >= 5):
        if len(homeGamesList[currentIndex]) == 0:
            currentIndex = currentIndex - 1
        else:
#           
#           get the team to switch
            
            switchHomeTeam = random.choice(homeGamesList[currentIndex])
#
#           now get the list of teams played
            teamsPlayed = cmlaDict[switchHomeTeam].getHomeGamesList()
            switchFound = False
            searchIndex = 0
            while switchFound == False:
                switchBucket = 0 
                gotList = False
                switchbucket = 0
                while (gotList == False):
                    if len(homeGamesList[switchBucket]) == 0:
                        switchBucket = switchBucket + 1
                    else:
                        switchList = homeGamesList[switchBucket]
                        intersectList = list(set(teamsPlayed).intersection(switchList))
                        if len(intersectList) == 0:
                            switchBucket = switchBucket + 1
                        else:
                            gotList = True
                if len(intersectList) == 0:
                    #print ('No team found to switch home/away', currentIndex, searchIndex)
                    switchFound = True
                else:
                    #print ('Found these teams to switch with',intersectList)
                    #
                    # pick the first team and do the book keeping
                    
                    switchAwayTeam = random.choice(intersectList)
                    switchBookkeeping(switchHomeTeam, switchAwayTeam, cmlaDict, scheduleTable)
                    homeGamesList[currentIndex].remove(switchHomeTeam)
                    homeGamesList[currentIndex - 1].append(switchHomeTeam)
                    homeGamesList[switchBucket].remove(switchAwayTeam)
                    homeGamesList[switchBucket + 1].append(switchAwayTeam)
                    switchFound = True
#            currentIndex = currentIndex - 1
                    

    return scheduleTable

def writeExcelInterimFile(masterSchedule):

    excelFilename = tkinter.filedialog.asksaveasfilename()
    
    keyList = list(masterSchedule.keys())
    keyList.sort()
    #
    # create new workbook
    wb = Workbook()

    for key in keyList:
        #
        # create sheet in file for key
        ws = wb.create_sheet()
        ws.title = key
        (scheduleTable, scheduleList, hasBye, numTeams) = masterSchedule[key]
        sortedList = []
        for team in scheduleList:
            if team != "BYE":
                sortedList.append(team)
        sortedList.sort()
        if hasBye == True:
            sortedList.append("BYE")
        
        ws.cell(row=0,column=0).value =  key
        cellrow = 1
        cellcol = 0
        for team in sortedList:
            ws.cell(row=cellrow,column=cellcol).value = team
            cellrow = cellrow + 1

        teamSchedule = []

        cellrow = 0
        cellcol = 3
        ws.cell(row=cellrow,column=cellcol).value = 'Vis'
        ws.cell(row=cellrow,column=cellcol+1).value = 'Hm'
        cellrow = 1
        for game in scheduleTable:
            (away,home) = game
            ws.cell(row=cellrow,column=cellcol).value = scheduleList[away]
            ws.cell(row=cellrow,column=cellcol+1).value = scheduleList[home]
            cellrow = cellrow + 1

        ws.cell(row=0,column=10).value = 'NOTES:'
        ws.cell(row=1,column=10).value = '1) COPY A WORKSHEET WITH '+ str(numTeams) + ' TEAMS AND RENAME TO ' + key
        ws.cell(row=2,column=10).value = '2) COPY TEAM LIST [COL A] TO [COL M] ON SHEET ' + key
        ws.cell(row=3,column=10).value = "3) COPY SCHEDULE TABLE [COL'S D & E] TO [COL'S D & E] ON SHEET " + key

    wb.save(excelFilename)

    return

def buildNotPlayedDict(notPlayed, depth):
    """
    Function to build dictionary for multiple parish teams not played list 

    Format will be 

    key = T[#]T[#] *** T[#]
    value = list of possible teams not played

    ex,

    T1T5 = [3,7,9,10]

    means that team 1 and team 5 don't play teams 3,7,9 or team 10

    OR

    T1T5T9 = [16]

    means teams 1, 5, and 9 don't play team 16

    """

    notPlayedDict = {}
    numTeams = len(notPlayed)

    if depth >= 3:
        for i in range(numTeams):
            for j in range(len(notPlayed[i])):
                try:
                    teamIndex = notPlayed[i][j]
                    if teamIndex != i:
                        intersectList = list(set(notPlayed[i]).intersection(notPlayed[teamIndex]))
                        if len(intersectList) > 0:
                            #
                            # first check if key exists from symmetry
                            try:
                                keyexists = "T" + str(teamIndex) + "T" + str(i)
                                if keyexists in notPlayedDict.keys():
                                    pass
                                else:
                                    key = "T" + str(i) + "T" + str(teamIndex)
                                    notPlayedDict[key] = intersectList
                            except:
                                pass
                except:
                    pass

    tokens = []
    if depth >= 4:
        tempDict = {}
        for key in notPlayedDict:
            teamList = notPlayedDict[key]
            for j in range(len(teamList)):
                try:
                    tokens = str(key).split("T")
                    tokens = list(filter(None, tokens))
                    teamIndex = teamList[j]
                    intersectList = list(set(teamList).intersection(notPlayed[teamIndex]))
                    if len(intersectList) > 0:
                        if (teamIndex) < int(tokens[0],10):
                            newKey = "T" + str(teamIndex) + key
                        elif teamIndex < int(tokens[1],10):
                            newKey = "T" + str(tokens[0]) + "T" + str(teamIndex) + "T" + str(tokens[1])
                        else:
                            newKey = key + "T" + str(teamIndex)
                        if newKey in tempDict:
                            pass
                        else:
                            tempDict[newKey] = intersectList
                except:
                    pass
        for key in tempDict:
            notPlayedDict[key] = tempDict[key]

    return notPlayedDict


def main():
    teamListFilename = ""
    registrationFilename = ""
    try:
        opts, args = getopt.getopt(sys.argv[1:], "hf:n:t:r:",["help","filename="])
    except getopt.error as msg:
        print (msg)
        print ("for help use --help")
        sys.exit(2)

    for o, arg in opts:
#        print (o, arg)
        if o == "-h":
            print ("python teammatrix.py [filename]")
        if o == "-f":
            filename = arg
        if o == "-n":
            numTeams = int(arg,10)
        if o == "-t":
            teamListFilename = arg
        if o == "-r":
            registrationFilename = arg

    if registrationFilename == "":
        registrationFilename = tkinter.filedialog.askopenfilename()

    tokens = registrationFilename.split('.')
    numTokens = len(tokens)
    extension = tokens[numTokens-1]

    if extension in ('xls','xlsx'):
        regDict = readExcelFile(registrationFilename)
    else:
        regDict = makeRegistrationDict(registrationFilename)
#       print ("Team list is ", teamListFilename)

    #
    # Now loop over all grade/genders to make the schedule
    rKeys = list(regDict.keys())
    rKeys.sort()
    masterSchedule = {}
    notPlayedDict = {}
    for rKey in rKeys:
        notPlayedDict.clear()
        if rKey[0] == "8":
            print (" Starting 8's")
        parishList = regDict[rKey]
        cmlaTeamDict, hasBye = makeGradeGenderTeamList(parishList)
        numTeams = len(cmlaTeamDict)
        scheduleTable = getScheduleTable(len(cmlaTeamDict))
        teamsPlayed = getTeamsPlayed(numTeams, scheduleTable)
        teamsNotPlayed = getTeamsNotPlayed(numTeams, teamsPlayed)
        notPlayedDict = buildNotPlayedDict(teamsNotPlayed, 4)
        scheduleList = makeSchedule(cmlaTeamDict, teamsNotPlayed, notPlayedDict)
        updateCMLADict(cmlaTeamDict, scheduleTable, scheduleList)
        scheduleTable = loadBalanceSchedule(cmlaTeamDict, scheduleTable, scheduleList)
        print('##########################')
        print('      ',rKey,' Schedule   ')
        print('##########################')
        tKeys = list(cmlaTeamDict.keys())
        tKeys.sort()
        for key in tKeys:
            print ('Team ', key, ' has ',cmlaTeamDict[key].getNumHomeGames(),'home games')
            print ('     ',cmlaTeamDict[key].getHomeGamesList())
            print ('  and ',cmlaTeamDict[key].getNumAwayGames(),'away games')
            print ('     ',cmlaTeamDict[key].getAwayGamesList())

        print()
        print()
        masterSchedule[rKey] = (scheduleTable, scheduleList, hasBye, len(scheduleList))
    writeExcelInterimFile(masterSchedule)
if __name__ == "__main__":
    main()




